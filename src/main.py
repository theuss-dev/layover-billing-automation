import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Importando nossos próprios módulos
from config import VALOR_SGL, VALOR_DBL, TAXA_ISS, VALOR_REFEICAO
from services import formatar_nome, calcular_refeicoes

def processar_faturamento():
    caminho_entrada = os.path.join('data', 'input', 'opera.csv')
    caminho_saida = os.path.join('data', 'output', 'faturamento_final.xlsx')

    if not os.path.exists(caminho_entrada):
        print(f"ERRO: O arquivo não foi encontrado em {caminho_entrada}")
        return

    print("Iniciando processamento com injeção de fórmulas do Excel...")
    quartos = {}

    with open(caminho_entrada, mode='r', encoding='utf-8') as file:
        reader = csv.DictReader(file, delimiter=';')
        
        for row in reader:
            if row['Reservation Type'] == 'Cancelled' or row['Room Type to Charge'] == 'PM' or not row['Room']:
                continue

            uh = row['Room']
            nome_formatado = formatar_nome(row['Name'])
            refeicoes_hospede = calcular_refeicoes(row['Arrival'], row['ETA'], row['Departure'], row['ETD'])
            
            checkin_formatado = datetime.strptime(row['Arrival'], "%d/%m/%Y").strftime("%d/%m/%Y")
            checkout_formatado = datetime.strptime(row['Departure'], "%d/%m/%Y").strftime("%d/%m/%Y")

            if uh not in quartos:
                quartos[uh] = {
                    'nomes': [], 'conf': int(row['Confirmation Number']),
                    'checkin': checkin_formatado, 'checkout': checkout_formatado,
                    'diarias': int(row['Nights']), 'total_refeicoes': 0
                }
            
            quartos[uh]['nomes'].append(nome_formatado)
            quartos[uh]['total_refeicoes'] += refeicoes_hospede

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturamento United"

    colunas_saida = [
        "NOME", "SGL OU DBL", "UH", "CONFIRMAÇÃO (INTERNA)", "CHECKIN", 
        "QUANT. DIÁRIAS", "CHECKOUT", "VALOR UNITÁRIO", "DIÁRIA + ISS", 
        "TOTAL", "ALIMENTAÇÃO UNITÁRIA", "Alimentações", 
        "VALOR TOTAL DE ALIMENTAÇÃO", "TOTAL DA ESTADIA"
    ]

    ws.append(colunas_saida)

    # --- DEFININDO ESTILOS ---
    borda_fina = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.border = borda_fina
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Inserindo os Dados com FÓRMULAS
    linha_atual = 2
    for uh, dados in quartos.items():
        tipo_quarto = "DBL" if len(dados['nomes']) > 1 else "SGL"
        nomes_juntos = ", ".join(dados['nomes'])
        valor_base = VALOR_DBL if tipo_quarto == "DBL" else VALOR_SGL

        ws.append([
            nomes_juntos, 
            tipo_quarto, 
            int(uh), 
            dados['conf'], 
            dados['checkin'],
            dados['diarias'], 
            dados['checkout'], 
            valor_base, 
            f"=H{linha_atual}*{TAXA_ISS}",                # Coluna I: Diária + ISS
            f"=I{linha_atual}*F{linha_atual}",           # Coluna J: Total de Diárias
            VALOR_REFEICAO, 
            dados['total_refeicoes'],
            f"=K{linha_atual}*L{linha_atual}",           # Coluna M: Total Alimentação
            f"=J{linha_atual}+M{linha_atual}"            # Coluna N: Total Estadia
        ])

        # Formatação das linhas normais
        for col_index in range(1, 15):
            celula = ws.cell(row=linha_atual, column=col_index)
            celula.border = borda_fina
            
            if col_index in [2, 3, 5, 6, 7, 12]:
                celula.alignment = Alignment(horizontal="center", vertical="center")
            elif col_index in [8, 9, 10, 11, 13, 14]:
                celula.number_format = '_-"R$" * #,##0.00_-;-"R$" * #,##0.00_-;_-"R$" * "-"??_-;_-@_-'
            
        linha_atual += 1

    # --- LINHA DE VALORES TOTAIS (Fórmulas SUM automáticas) ---
    linha_total = [
        "VALORES TOTAIS", "", "", "", "", "", "", "", "", 
        f"=SUM(J2:J{linha_atual-1})",  # Soma total de diárias
        "", 
        f"=SUM(L2:L{linha_atual-1})",  # Soma total de refeições físicas
        f"=SUM(M2:M{linha_atual-1})",  # Soma total financeira de alimentação
        f"=SUM(N2:N{linha_atual-1})"   # Soma do Faturamento Final
    ]
    ws.append(linha_total)
    ultima_linha = ws.max_row

    ws.merge_cells(start_row=ultima_linha, start_column=1, end_row=ultima_linha, end_column=9)

    # Estilizando a linha de Totais
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    total_font = Font(bold=True)

    for col_num in range(1, 15):
        cell = ws.cell(row=ultima_linha, column=col_num)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = borda_fina
        
        if col_num == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_num in [10, 13, 14]:
            cell.number_format = '_-"R$" * #,##0.00_-;-"R$" * #,##0.00_-;_-"R$" * "-"??_-;_-@_-'
        elif col_num == 12:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuste Automático da Largura das Colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 3)
        ws.column_dimensions[col_letter].width = adjusted_width

    ws.auto_filter.ref = f"A1:N{ultima_linha - 1}"

    wb.save(caminho_saida)
    print(f"✅ SUCESSO! A planilha corporativa COM FÓRMULAS está pronta em: {caminho_saida}")

if __name__ == "__main__":
    processar_faturamento()